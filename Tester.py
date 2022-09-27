from os.path import exists

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

import BST
import RBT
import random


def fillTree(tree, n, rand):
    values = []
    for x in range(0, n, 1):
        values.append(x)
    if rand:
        random.shuffle(values)
    for x in values:
        tree.insert(x)
    return tree


class Tester:
    def __init__(self):
        self.header = ["ID", "N.nodi", "Altezza"]
        self.seqRelativePath = "out/sequential.xlsx"
        self.randRelativePath = "out/random.xlsx"
        self.BST_CODE = "BST"
        self.RBT_CODE = "RBT"

    def runTest(self, path, n, rand):
        if not exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Output"
            ws.append(self.header)
            wb.save(path)

        wb = load_workbook(path)
        ws = wb.active

        rbTree = RBT.RBT()
        binTree = BST.BST()

        rbTree = fillTree(rbTree, n, rand)
        binTree = fillTree(binTree, n, rand)

        binH = binTree.root.height()
        rbH = rbTree.root.height()

        ws.append([self.BST_CODE, n, binH])
        ws.append([self.RBT_CODE, n, rbH])
        wb.save(path)

